# coding:utf-8
import os
import path_config
from utils.Handlelogger import HandleLogging
from openpyxl import Workbook
from openpyxl.styles import Font, colors

# 初始化计数器，检测python环境是否安装第三方库openpyxl
i = 2
while i:
    try:
        from openpyxl import load_workbook

        # logging.info("已安装外部第三方库")
        break

    except:
        os.system("pip3 install openpyxl")
        i -= 1
        continue


class CasesData(object):
    '''用来保存用例数据的类'''
    pass


class HandleExcel(object):
    '''读取excel工具类。'''

    # 行和列都是从1开始数起
    def __init__(self, filename, sheetname=None):
        '''实例化文件属性，初始化操作文件对象'''
        self.filename = filename
        self.sheetname = sheetname

        # 定义时间格式和日志类容输出格式
        self.log = HandleLogging(
            logger=os.path.basename(__file__).split(".")[0]).getlog()

    #         创建打开文件对象,在构造函数中初始化，在执行测试用例，结果写入时，有多个sheet结果写入，只会在执行最后一个sheet有写入结果
    #         self.wb=load_workbook(self.filename)
    #         默认sheetname为None，传第一个sheet
    #         self.ws=self.wb[self.sheetname] if self.sheetname is not None else self.wb.active

    def open(self):
        '''打开工作簿及表单对象'''
        try:
            #         创建工作簿对象
            self.wb = load_workbook(self.filename)
            #         默认sheetname为None，传第一个sheet
            self.ws = self.wb[
                self.sheetname] if self.sheetname is not None else self.wb.active
        except Exception as e:
            self.log.error("{}文件不存在".format(self.filename))
            raise e

    def close(self):
        '''关闭工作簿对象'''
        self.wb.close()

    def get_all_cases(self):
        '''获取excel所有行的测试用例,这个所谓所有,即需要添加对应的列column'''
        """读取数据"""
        self.open()
        # 获取最大的行
        max_row = self.ws.max_row
#         应该有最大列,在下面的循环使用读取对应的行和列
#         max_col = self.ws.max_column
        # 读取所有的数据，放到一个列表中
        list_data = []
        for i in range(1, max_row + 1):
            data1 = self.ws.cell(row=i, column=1).value
            data2 = self.ws.cell(row=i, column=2).value
            data3 = self.ws.cell(row=i, column=3).value
            data4 = self.ws.cell(row=i, column=4).value
            list_data.append([data1, data2, data3, data4])

        self.close()
        # 创建一个字典，用来存储所有的用来数据
        cases = []
        # 获取第一行数据中的表头
        title = list_data[0]
        for data in list_data[1:]:
            # 遍历第一行之外的其他行数据，聚合打包成字典
            case = dict(zip(title, data))
            cases.append(case)

        return cases

    def get_datas(self):
        '''获取excel所有行的测试用例'''
        """读取数据较上一个方法更灵活"""
        self.open()
        # 获取表单所有行
        rows = list(self.ws.rows)
        # 读取工作簿表单所有行数据之后，就可以关闭
        self.close()
        # 接收表头的空列表
        title = []
        for t in rows[0]:  # 0,第一行就是title
            title.append(t.value)

        # 接收所有测试用例数据的空列表
        cases = []
        for row in rows[1:]:  # 截取list表头后面的元素，遍历
            data = []
            for r in row:
                # 遍历每一行的单元格的值加入一个空列表
                data.append(r.value)
                # zip打包函数
            # 一行测试用例数据，数据格式是dict的items类型，可以转成dict
            case = dict(zip(title, data))
            # 每一行测试用例数据追加成一个用例集合
            cases.append(case)
        return cases

    def get_datas_obj(self):
        '''获取excel所有行的测试用例'''
        """读取数据"""
        self.open()
        # 获取表单所有行
        rows = list(self.ws.rows)
        # 关闭工作簿对象
        self.close()
        # 接收表头的空列表
        title = []
        for t in rows[0]:  # 0,第一行就是title
            title.append(t.value)
        # 定义空列表用例集合
        cases = []
        for row in rows[1:]:  # 截取list表头后面的元素，遍历
            data = []
            for r in row:
                # 遍历每一行的单元格的值加入一个空列表
                data.append(r.value)
                # zip打包函数
            case = list(zip(title, data))  # 一行测试用例数据，元组元素组成的列表数据
            # 创建对象用来保存一行的用例数据
            case_obj = CasesData()
            # 遍历列表中该行的元素，是个元组，k，v元组拆包
            for k, v in case:
                setattr(case_obj, k, v)  # 设置对象的属性及属性值
            # 追加一行测试对象到用例集合列表
            cases.append(case_obj)

        return cases

    def get_one_case(self, row):
        '''获取指定行的测试用例数据，row为case_id顺位'''
        # 打开工作簿对象
        self.open()
        # 获取最大的列和行，最小行
        max_col = self.ws.max_column
        max_row = self.ws.max_row
        min_row = self.ws.min_row
        # 获取所有行
        rows = list(self.ws.rows)
        # 关闭工作簿对象
        self.close()

        title = []  # 表头
        for row in rows[0]:
            title.append(row.value)

        list_data = []
        #         先判断行在范围内
        if isinstance(row, int) and min_row <= row <= max_row:
            for j in range(1, max_col + 1):
                # 提取每个单元格的数据，重新组成一个列表
                data1 = self.ws.cell(row=row + 1, column=j).value
                print(data1)
                list_data.append(data1)
            case = dict(zip(title, list_data))
            return case
        else:
            self.log.error("你输入的行号:{}不正确!".format(str(row)))

    def get_title_col_index(self, kword):
        '''获取表头关键字参数，返回索引，便于结合行列准确写入结果
                    用于python37版本后，有序的list
        '''
        # 打开
        self.open()
        #         获取所有行
        rows = list(self.ws.rows)
        # 关闭
        self.close()

        title = []  # 表头
        for row in rows[0]:
            title.append(row.value)

        if kword in title:
            return title.index(kword) + 1
        else:
            self.log.error("title中不存在该表头：{}".format(kword))

            # 虽然可以少一些代码，但是运算速度会低一些
            #         for i in range(len(rows[0])):
            #             #遍历表头，返回列的索引
            #             if rows[0][i].value == kword:
            #                 return i+1

    def get_data_row_index(self, data):
        '''获取执行用例所在行，用于回写结果于哪一行
                    用于python37版本后，有序的list
        '''
        #         获取所有测试用例数据
        datas = self.get_datas()
        #         再查出元素所在列表的索引，
        if data in datas:
            # 获取行的索引是从0开始，第0行是title，在读取测试用例数据的行本身就是从1开始，所以+2
            return datas.index(data) + 2
        else:
            self.log.error('datas没有该行数据:{}'.format(data))

    def write_data(self, data, kword, result_status):
        '''执行测试用例，根据用例所在行，写入结果列，并保存
        data:读取excel遍历的测试数据case：应该是按索引取一行测试用例数据，那么row-1
        kword:读取excel表头的索引，表头关键字：即为dict的[key]
        message:写入excel的内容
        return:
                     注意:这个方法因为要先调用两个方法获取行跟列,再调用写入方法,从效率上来讲会比较慢
        '''
        row = self.get_data_row_index(data)
        column = self.get_title_col_index(kword)

        if isinstance(row, int) and isinstance(column, int):
            self.write_file(row, column, result_status)
        else:
            self.log.info("获取的行:{}或者列:{}不是数字".format(row, column))

    def write_file(self, row, column, result_status):
        '''执行用例结果写入excel，并保存
        row:行
        column:列
        result_status:写入excel的列'''
        if isinstance(row, int) and isinstance(column, int):
            try:
                self.open()
                # self.get_col_kw(kword)
                self.ws.cell(row=row, column=column, value=result_status)
            except Exception as e:
                self.log.info("写入异常")
                raise e
            else:
                self.wb.save(self.filename)
                self.log.info("写入成功")
            finally:
                self.close()
        else:
            self.log.info("输入的行:{}或者列:{}不是数字".format(row, column))


class Write_excel(object):
    """修改excel数据"""

    def __init__(self, filename):
        '''初始化文件对象'''
        self.filename = filename
#         创建xlsx文件,如果不存在,顺便写上头
        if not os.path.exists(self.filename):
            self.wb = Workbook()
            self.ws = self.wb.active  # 激活sheet
            self.ws.cell(1, 1).value = "caseid"
            self.ws.cell(1, 2).value = "title"
            self.ws.cell(1, 3).value = "env"
            self.ws.cell(1, 4).value = "path"
            self.ws.cell(1, 5).value = "method"
            self.ws.cell(1, 6).value = "params"
            self.wb.save(filename)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active

    def write(self, row_n, col_n, value):
        """写入数据，如(2,3，"hello"),第二行第三列写入数据"hello\""""
        ft = Font(color=colors.RED, size=12, bold=True)
        # 判断值为错误时添加字体样式
        if value in ['fail', 'error'] or col_n == 12:
            self.ws.cell(row_n, col_n).font = ft
        if value == 'pass':
            ft = Font(color=colors.GREEN)
            self.ws.cell(row_n, col_n).font = ft

        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)


if __name__ == '__main__':
    excel = HandleExcel(path_config.account_excel)
    datas = excel.get_all_cases()
    print(datas)
#     i = random.randint(0, len(datas) - 1)
#     data = datas[i]
#     row = excel.get_data_row_index(data)
#     print(data['account'])
#
#     exc = HandleExcel(path_config.account_excel)
#     exam_date = exc.get_datas()
#     print(exam_date[0]['account'])
    # print(row,data)
    # excel.write_data(data, "result", "通过")
    #     for r in excel.get_datas():
    #         print(r.__dict__)
    #     excel.write_data(3, "result", "通过")
    #     print(excel.get_title_index("result"))
