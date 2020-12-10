'''
Created on 2020年7月10日

@author: qguan
'''
import json
import os

from openpyxl import Workbook

import path_config


class Write_excel(object):
    """修改excel数据"""

    def __init__(self, filename):
        '''初始化文件对象'''
        self.filename = filename
#         创建xlsx文件,如果不存在,顺便写上头
        if not os.path.exists(self.filename):
            self.wb = Workbook()
            self.ws = self.wb.active  # 激活sheet
            self.ws.cell(1, 1).value = "url"
            self.ws.cell(1, 2).value = "params"
            self.wb.save(filename)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active

    def write(self, row_n, col_n, value):
        """写入数据，如(2,3，"hello"),第二行第三列写入数据"hello\""""
        self.ws.cell(row_n, col_n).value = value

    def save(self):
        self.wb.save(self.filename)


def get_datas(file):
    with open(file=file) as pf:
        text = json.loads(pf.read())
        testcases = text["teststeps"]
        datas_list = []
        for case in testcases:
            datas = {}
            datas['url'] = case["request"]["url"]
            if case["request"].get('json'):
                datas['params'] = case["request"]["json"]
            datas_list.append(datas)
    return datas_list

if __name__ == '__main__':
    excel = Write_excel('testcase.xlsx')
    datas = get_datas(path_config.jsonData_path + 'lesson_push_o_after.json')
    for i in range(len(datas)):
        url = datas[i]['url']
        if datas[i].get('params'):
            excel.write(i + 2, 1, url)
            excel.write(i + 2, 2, str(datas[i].get('params')))
        else:
            excel.write(i + 2, 1, url)
    excel.save()
